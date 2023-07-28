import ast
import datetime
from tempfile import NamedTemporaryFile
from typing import Dict

from django.core.exceptions import ValidationError, ImproperlyConfigured
from django.db.models import QuerySet
from django.http import QueryDict
from django.utils.formats import date_format
from django.utils.translation import gettext as _, gettext_lazy
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from base.models.person import Person


class ExportMixin:
    def get_export_objects(self, **kwargs):
        raise NotImplementedError


class QuerySetExportMixin(ExportMixin):
    def get_export_objects(self, **kwargs):
        return self.get_queryset_export(kwargs.get("filters"))

    def get_queryset_export(self, filters) -> QuerySet:
        raise NotImplementedError


class FilterSetExportMixin(QuerySetExportMixin):
    def get_queryset_export(self, filters) -> QuerySet:
        # Generate a dict from the saved querydict
        filters_dict = QueryDict(filters).dict()
        # Now get the filterset class and instantiate it with the filters
        filterset = self.get_filterset_class()(data=filters_dict)
        if filterset.is_valid():
            # and finally get back the initial queryset results
            return filterset.qs.all()
        raise ValidationError("The formset is not valid")


class FileExportMixin:
    file_extension = None
    mimetype = None

    def generate_file(self, person, filters):
        raise NotImplementedError

    def get_mimetype(self):
        if self.mimetype is None:
            raise ImproperlyConfigured("Specify mimetype on mixin class")
        return self.mimetype

    def get_file_extension(self):
        if self.file_extension is None:
            raise ImproperlyConfigured("Specify file_extension on mixin class")
        return self.file_extension

    def get_read_token_extra_kwargs(self) -> Dict:
        return {}

    def get_task_started_async_manager_extra_kwargs(self) -> Dict:
        return {}

    def get_task_error_async_manager_extra_kwargs(self, e: Exception) -> Dict:
        return {}

    def get_task_done_async_manager_extra_kwargs(self, file_name: str, file_url: str) -> Dict:
        return {}


class ExcelFileExportMixin(FileExportMixin):
    file_extension = ".xlsx"
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    with_legend_worksheet = False
    with_parameters_worksheet = False
    description = gettext_lazy('List')

    def get_header(self):
        raise NotImplementedError

    def get_row_data(self, row):
        raise NotImplementedError

    def customize_legend_worksheet(self, worksheet: Worksheet):
        """Override it to customize the legend worksheet"""
        pass

    def get_formatted_filters_parameters_worksheet(self, filters: str) -> Dict:
        """Override it to customize the configuration filters. Must return the formatted filters as dict"""
        try:
            return ast.literal_eval(filters)
        except ValueError:
            return {}

    def customize_parameters_worksheet(self, worksheet: Worksheet, person: Person, filters: str):
        """Override it to customize the configuration worksheet"""
        worksheet.append([_('Creation date'), date_format(datetime.date.today())])
        worksheet.append([_('Created by'), person.full_name if person else _('Unknown')])
        worksheet.append([_('Description'), f"{_('Export')} - {self.description}"])

        if filters:
            formatted_filters = self.get_formatted_filters_parameters_worksheet(filters)
            for key, value in formatted_filters.items():
                worksheet.append([key, str(value)])

    def customize_workbook_before_save(self, workbook: Workbook):
        """Override it to customize the created workbook"""
        pass

    def generate_file(self, person, filters, **kwargs):
        workbook = Workbook()

        # grab the active worksheet
        worksheet = workbook.active

        worksheet.title = str(self.description)

        # add headers
        worksheet.append(self.get_header())
        cells = worksheet.iter_rows(min_row=1, max_row=1)
        for col in cells:
            for cell in col:
                cell.font = Font(bold=True)

        # add data
        export_objects = self.get_export_objects(filters=filters)
        for export in export_objects:
            worksheet.append(self.get_row_data(export))

        # add legend
        if self.with_legend_worksheet:
            legend_worksheet = workbook.create_sheet(title=_('Legend'))
            self.customize_legend_worksheet(worksheet=legend_worksheet)

        # add configuration
        if self.with_parameters_worksheet:
            configuration_worksheet = workbook.create_sheet(title=_('Parameters'))
            self.customize_parameters_worksheet(
                worksheet=configuration_worksheet,
                person=person,
                filters=filters,
            )

        self.customize_workbook_before_save(workbook)

        # stream back the file
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            return tmp.read()


class ExcelFilterSetExportMixin(FilterSetExportMixin, ExcelFileExportMixin):
    """Excel export from a FilterSet based view"""
