import uuid
from unittest.mock import patch

from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.core.files.base import ContentFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl.styles import Font

from base.tests.factories.person import PersonFactory
from osis_export.contrib.export_mixins import (
    ExcelFileExportMixin,
    ExportMixin,
    FileExportMixin,
)
from osis_export.models import Export
from osis_export.tests.export_test.models import DummyModel
from osis_export.tests.export_test.views import TestViewSearch


@override_settings(
    OSIS_EXPORT_ASYNCHRONOUS_MANAGER_CLS='osis_export.tests.export_test.async_manager.AsyncTaskManager',
)
class TestAsyncExport(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.person = PersonFactory()
        cls.export_data = {
            "async_task_name": "test name",
            "async_task_description": "test description",
            "called_from_class": "osis_export.tests.export_test.views.TestViewSearch",
            "filters": "",
            "type": "EXCEL",
            "file_name": "file test name",
            "next": "/some-url",
        }
        cls.url = reverse("osis_export:export")

    def setUp(self) -> None:
        self.client.force_login(self.person.user)

    def test_async_export_view_returns_errors_message_if_no_data_given(self):
        # send a post with no data must raise an exception as their is required fields
        url = self.url
        response = self.client.post(url, {})
        self.assertEqual(response.status_code, 302)
        response = self.client.post(url, {}, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertIsNot(len(list(response.context["messages"])), 0)
        self.assertIn(
            "called_from_class",
            list(response.context["messages"])[0].message,
        )

    @patch('osis_export.tests.export_test.async_manager.AsyncTaskManager.create')
    def test_async_export_view_must_redirect_to_next_parameter_if_valid(self, create_task):
        create_task.return_value = uuid.uuid4()
        response = self.client.post(self.url, self.export_data)
        self.assertEqual(response.status_code, 302)
        # must redirect to the given 'next' parameter
        self.assertEqual(response.url, self.export_data["next"])

    def test_async_export_view_must_redirect_to_slash_if_valid_and_next_not_set(self):
        data = self.export_data.copy()
        del data["next"]
        response = self.client.post(self.url, data)
        self.assertEqual(response.status_code, 302)
        # must redirect to the given 'next' parameter
        self.assertEqual(response.url, "/")

    def test_async_export_view_must_redirect_to_next_parameter_if_invalid(self):
        response = self.client.post(self.url, {"next": self.export_data["next"]})
        self.assertEqual(response.status_code, 302)
        # must redirect to the given 'next' parameter
        self.assertEqual(response.url, self.export_data["next"])

    def test_async_export_view_must_redirect_to_slash_if_invalid_and_next_not_set(self):
        response = self.client.post(self.url, {})
        self.assertEqual(response.status_code, 302)
        # must redirect to the given 'next' parameter
        self.assertEqual(response.url, "/")

    @patch('osis_export.tests.export_test.async_manager.AsyncTaskManager.create')
    def test_async_export_view_set_default_ttl_if_not_given(self, create_task):
        create_task.return_value = uuid.uuid4()
        self.assertEqual(Export.objects.count(), 0)
        create_task.assert_not_called()

        response = self.client.post(self.url, self.export_data)
        self.assertEqual(response.status_code, 302)

        # Export must have been created and create() called
        self.assertEqual(Export.objects.count(), 1)
        create_task.assert_called_with(
            name=self.export_data["async_task_name"],
            description=self.export_data["async_task_description"],
            person=self.person,
        )

        # now test the values of the related export
        export = Export.objects.get()
        self.assertEqual(
            export.called_from_class, self.export_data["called_from_class"]
        )
        self.assertEqual(export.filters, self.export_data["filters"])
        self.assertEqual(export.person, self.person)
        self.assertEqual(export.job_uuid, create_task.return_value)
        self.assertEqual(export.file, [])
        self.assertEqual(export.file_name, self.export_data["file_name"])
        self.assertEqual(export.type, self.export_data["type"])
        self.assertIsNotNone(export.created_at)

    @patch('osis_export.tests.export_test.async_manager.AsyncTaskManager.create')
    def test_async_export_view_with_ttl(self, create_task):
        create_task.return_value = uuid.uuid4()
        self.assertEqual(Export.objects.count(), 0)

        data = {**self.export_data, "async_task_ttl": 42}
        response = self.client.post(self.url, data)
        self.assertEqual(response.status_code, 302)

        # Export must have been created and create() called
        self.assertEqual(Export.objects.count(), 1)
        create_task.assert_called_with(
            name=self.export_data["async_task_name"],
            description=self.export_data["async_task_description"],
            person=self.person,
            time_to_live=42,
        )


class TestFileExportMixin(TestCase):
    @classmethod
    def setUpTestData(cls):
        class MyClass(FileExportMixin):
            pass

        cls.my_instance = MyClass()
        cls.MyClass = MyClass

    def test_file_export_mixin_must_specify_mimetype(self):
        with self.assertRaises(ImproperlyConfigured):
            self.my_instance.get_mimetype()

        self.MyClass.mimetype = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertEqual(self.my_instance.get_mimetype(), self.MyClass.mimetype)

    def test_file_export_mixin_must_specify_file_extension(self):
        with self.assertRaises(ImproperlyConfigured):
            self.my_instance.get_file_extension()

        self.MyClass.file_extension = ".xlsx"
        self.assertEqual(
            self.my_instance.get_file_extension(),
            self.MyClass.file_extension,
        )


class TestExcelFileExportMixin(TestCase):
    @classmethod
    def setUpTestData(cls):
        class MyExportableObject:
            test_param = "test"
            test_param_2 = None
            test_param_3 = "This value should not appear on Excel export"
            test_param_4 = "Some value"

        cls.my_object = MyExportableObject()

        class MyClass(ExportMixin, ExcelFileExportMixin):
            def get_export_objects(self, **kwargs):
                return [MyExportableObject() for _ in range(10)]

            def get_header(self):
                return ["test param", "test param 2", "test param 4"]

            def get_row_data(self, row):
                return [row.test_param, row.test_param_2, row.test_param_4]

        cls.my_class_instance = MyClass()

    def test_generate_file_creates_excel_file(self):
        file = self.my_class_instance.generate_file()
        self.assertEqual(type(file), bytes)
        workbook = load_workbook(ContentFile(file))
        worksheet = workbook.active
        # check if the first row is in bold style
        cells = worksheet.iter_rows("A1:C1")
        # check that we have 11 rows : 1 for the header and 10 for data
        self.assertEqual(worksheet.get_highest_row(), 11)
        for col in cells:
            for cell in col:
                self.assertEqual(cell.font, Font(bold=True))
        # check if the workbook contains the correct values
        cells = worksheet.iter_rows("A2")
        for col in cells:
            for cell in col:
                self.assertEqual(cell.value, self.my_object.test_param)
        cells = worksheet.iter_rows("B2")
        for col in cells:
            for cell in col:
                self.assertEqual(cell.value, self.my_object.test_param_2)
        cells = worksheet.iter_rows("C2")
        for col in cells:
            for cell in col:
                self.assertEqual(cell.value, self.my_object.test_param_4)


class TestFilterSetExportMixin(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.dummy_objects_count = 10
        for cpt in range(cls.dummy_objects_count):
            DummyModel.objects.create(
                name="dummy-name-{}".format(cpt), selectable_value="A"
            )

        cls.my_class_instance = TestViewSearch()

    def test_return_all_objects_if_no_filters_specified(self):
        qs = self.my_class_instance.get_queryset_export("")
        self.assertEqual(qs.count(), self.dummy_objects_count)
        qs = self.my_class_instance.get_export_objects(filters="")
        self.assertEqual(qs.count(), self.dummy_objects_count)

    def test_return_filtered_queryset(self):
        qs = self.my_class_instance.get_queryset_export("name=dummy-name-2")
        self.assertEqual(qs.count(), 1)
        qs = self.my_class_instance.get_export_objects(filters="name=dummy-name-2")
        self.assertEqual(qs.count(), 1)

    def test_raises_validation_error_if_filterset_is_not_valid(self):
        with self.assertRaises(ValidationError):
            self.my_class_instance.get_queryset_export("selectable_value=Z")
